import pandas as pd
import numpy as np
import os

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

from bs4 import BeautifulSoup

from google import genai
from dotenv import load_dotenv

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

import platform
import subprocess


def load_dfs(import_path):
    import_df = pd.read_excel(import_path)

    columns = [
        "Patent Number",
        "Google Patent Link",
        "Patent Title",
        "Filing Date",
        "Expiry Date",
        "Abstract",
        "Claim 1",
        "Patent Type",
        "Main Topic/Subject of the invention",
        "Patent Review",
    ]
    export_df = pd.DataFrame(columns=columns)

    return import_df, export_df


def clean_dfs():
    # clean import_df
    import_df["Patent Number"] = import_df["Patent Number"].astype(str)

    # clean export_df
    # export_df = export_df.astype(str)


def simple_html_search(
    patent_number, search_term, soup, html_element, html_attributes, log_func=print
):

    searched_term = ""
    try:
        searched_term = soup.find(html_element, attrs=html_attributes).text.strip()
    except AttributeError:
        log_func(f"{patent_number}: {search_term} not found")

    return searched_term


def prompt_gemini(content_list):

    response = client.models.generate_content(
        model="gemini-2.0-flash-lite", contents=content_list
    )
    return response.text


def add_rows(us_patents, log_func=print):
    driver = webdriver.Firefox()

    for index, row in import_df.iterrows():
        # go through every row
        patent_number = row["Patent Number"]
        if us_patents:
            if not patent_number[0:2].isalpha():
                patent_number = "US" + patent_number

        driver.get("https://patents.google.com/")
        time.sleep(2)

        search = driver.find_element(By.NAME, "q")
        search.clear()
        search.send_keys(patent_number)
        search.send_keys(Keys.RETURN)

        time.sleep(2)
        patent_url = driver.current_url

        log_func(f"{patent_number}: Processing page")

        if "patents.google.com/patent" in patent_url:
            # do the rest of the stuff
            html = driver.page_source
            soup = BeautifulSoup(html, "html.parser")

            # Patent Title
            patent_title = simple_html_search(
                patent_number, "Title", soup, "h1", {"id": "title"}, log_func=log_func
            )

            # Filing Date
            filing_date = ""
            try:
                filing_date = soup.find(
                    "div", attrs={"class": "priority style-scope application-timeline"}
                ).text.strip()
            except AttributeError:
                try:
                    filing_date = soup.find(
                        "div", attrs={"class": "filed style-scope application-timeline"}
                    ).text.strip()
                except AttributeError:
                    log_func(f"{patent_number}: Filing Date not found")

            # Expiry Date
            expiry_date = ""
            try:
                legal_status_texts = soup.find_all(
                    "div", class_="legal-status style-scope application-timeline"
                )
                if legal_status_texts:
                    expiry_date = legal_status_texts[1].text.strip()
                else:
                    log_func(f"{patent_number}: Expiry Date not found (empty list)")
            except:
                log_func(f"{patent_number}: Expiry Date not found (index error)")

            # Abstract
            abstract = simple_html_search(
                patent_number,
                "Abstract",
                soup,
                "div",
                {"class": "abstract style-scope patent-text"},
                log_func=log_func,
            )

            # Claims text
            # Used for: Claim 1 and Patent Review
            claims_text = None
            claims_text = simple_html_search(
                patent_number,
                "Claims",
                soup,
                "patent-text",
                {"name": "claims", "class": "style-scope patent-result"},
                log_func=log_func,
            )

            # Description Text
            description_text = None
            description_text = simple_html_search(
                patent_number,
                "Description",
                soup,
                "patent-text",
                {
                    "name": "description",
                    "id": "descriptionText",
                    "class": "style-scope patent-result",
                },
                log_func=log_func,
            )

            # Claim 1
            # Extract Claim 1 from claims_text
            if claims_text:

                claim1 = prompt_gemini(
                    [
                        claims_text,
                        "Extract the first claim from this text. Do not use markdown formatting.",
                    ]
                )
            else:
                claim1 = None

            # Patent Review
            # Using claims_text, summarize each claim
            """
            patent_review = prompt_gemini(
                [
                    claims_text,
                    "Summarize each claim in a few sentences each. Do not use markdown formatting.",
                ]
            )"""

            # Patent Type
            # Using description_text and claims_text, choose between circuit, process, package, and system
            # circuit: Integrated circuits are compact electronic chips made up of interconnected components that include resistors, transistors, and capacitors.
            # process: Process of creating an electronic device
            # package: A semiconductor package is a metal, plastic, glass, or ceramic casing containing one or more discrete semiconductor devices or integrated circuits
            # system: If it's not the other three, put it into system.
            if claims_text and description_text:

                patent_type = prompt_gemini(
                    [
                        description_text,
                        claims_text,
                        "Using information from the patent description and patent claims, categorize the patent into one of the following categories: "
                        "Circuit, Process, Package, or System. "
                        "Use the following definitions to categorize the patents:"
                        "Circuit: Integrated circuits are compact electronic chips made up of interconnected components that include resistors, transistors, and capacitors."
                        "Process: The process of creating or manufacturing an electronic device."
                        "Package: A semiconductor package is a metal, plastic, glass, or ceramic casing containing one or more discrete semiconductor devices or integrated circuits."
                        "System: If it's not the other three, classify the patent as System"
                        "Do not use markdown formatting.",
                    ]
                )
            else:
                patent_type = None

            # Main Topic/Subject of the Invention
            # Using description_text, summarize the main topic of the invention in a few sentences
            """
            main_topic = prompt_gemini(
                [
                    description_text,
                    "In a few sentences, summarize the main topic of the invention. Do not use markdown formatting.",
                ]
            )
            """

            # add a new row
            new_row = {
                "Patent Number": patent_number,
                "Google Patent Link": patent_url,
                "Patent Title": patent_title,
                "Filing Date": filing_date,
                "Expiry Date": expiry_date,
                "Abstract": abstract,
                "Claim 1": claim1,
                "Patent Type": patent_type,
                # "Main Topic/Subject of the invention": main_topic,
                # "Patent Review": patent_review,
            }

            # limited to 30 requests per minute
            # 4 requests per patent
            # 7 patents/min. Sleep for 9 sec
            time.sleep(5)

        elif "patents.google.com/?" in patent_url:
            log_func(
                f"{patent_number}: multiple patents found, please check the patent number afterwards"
            )
            new_row = {
                "Patent Number": patent_number,
                "Google Patent Link": patent_url,
                "Patent Title": "Multiple patents found, please check the patent number",
            }

        else:
            log_func(f"{patent_number}: was not found")
            new_row = {
                "Patent Number": patent_number,
                "Google Title": "Patent not found",
            }

        export_df.loc[len(export_df)] = new_row
        time.sleep(2)

    driver.quit()


def save_file():
    # save the export file
    with pd.ExcelWriter(export_path) as writer:
        # import_df.to_excel(writer, sheet_name='Import', index = False)
        export_df.to_excel(writer, sheet_name="Export", index=False)


def format_saved_file():
    wb = load_workbook(export_path)
    ws = wb.active

    # add hyperlinks
    # Start from row 2 (since row 1 is headers)
    for row in range(2, len(export_df) + 2):
        link = ws[f"B{row}"].value
        if link:
            ws[f"B{row}"].value = ws[f"A{row}"].value
            ws[f"B{row}"].hyperlink = link
            ws[f"B{row}"].style = "Hyperlink"

    # stuff with header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
        cell.fill = PatternFill(
            start_color="8ED973", end_color="8ED973", fill_type="solid"
        )

    # change column dimensions:
    sheet = wb["Export"]
    sheet.column_dimensions["A"].width = 20  # patent number
    sheet.column_dimensions["B"].width = 20  # Google Patent Link
    sheet.column_dimensions["C"].width = 20  # Patent Title
    sheet.column_dimensions["D"].width = 12  # Filing Date
    sheet.column_dimensions["E"].width = 12  # Expiry Date
    sheet.column_dimensions["F"].width = 25  # Abstract
    sheet.column_dimensions["G"].width = 50  # Claim 1
    sheet.column_dimensions["H"].width = 12  # Patent Type
    sheet.column_dimensions["I"].width = 25  # Main Topic/Subject of the invention
    sheet.column_dimensions["J"].width = 35  # Patent Review

    # freeze the header and patent number
    ws.freeze_panes = "B2"

    # save it again at the end
    wb.save(export_path)


def open_file_location(filepath):
    if platform.system() == "Windows":
        # Show folder and highlight file
        os.startfile(os.path.normpath(filepath), "open")
    elif platform.system() == "Darwin":
        # macOS: open Finder at the folder
        subprocess.run(["open", "-R", filepath])
    else:
        # Linux: open the folder with default file manager
        subprocess.run(["xdg-open", os.path.dirname(filepath)])


def run_patent_scraper(
    import_file_path, export_file_path, us_patents, gemini_key, log_func=print
):
    global import_df, export_df, export_path, client

    export_path = export_file_path
    import_path = import_file_path

    client = genai.Client(api_key=gemini_key)

    import_df, export_df = load_dfs(import_path)
    clean_dfs()
    add_rows(us_patents, log_func)
    save_file()
    format_saved_file()
    open_file_location(export_path)
